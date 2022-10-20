import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from copy import copy


def is_number(s):
    try:
        float(s)
        return True
    except:
        return False

def highlight_cells(path):
    '''Note: This is the customized code, not for general use'''
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    yellow = "00FFFF00"

    grp0_row, grp1_row, grp2_row, last_row = 0, 0, 0, 0
    for i in range(max_row):
        if sheet[f'A{i+1}'].value and sheet[f'A{i+1}'].value.strip() == 'Group':
            grp0_row = i + 1
        elif sheet[f'A{i+1}'].value and sheet[f'A{i+1}'].value.strip() == 'Income':
            grp1_row = i + 1
        elif sheet[f'A{i+1}'].value and sheet[f'A{i+1}'].value.strip() == 'Expense':
            grp2_row = i + 1
        elif sheet[f'A{i+1}'].value and sheet[f'A{i+1}'].value.strip() == 'NOI':
            last_row = i + 1
            break
    cols = []
    for row in sheet.iter_rows(min_row=grp0_row, max_row=grp0_row, min_col=1, max_col=max_column):
        for cell in row:
            if cell.value and cell.value.strip() == 'Delta %':
                cols.append(cell.column)

    # Highlight cells in "Income" group
    for row in sheet.iter_rows(min_row=grp1_row, max_row=grp2_row-1, min_col=cols[0], max_col=cols[-1]):
        for cell in row:
            if is_number(cell.value) and cell.column in cols and cell.value <= -0.05:
                cell.fill = PatternFill(
                    start_color=yellow, end_color=yellow, fill_type='solid')

    # Highlight cells in "Expense" group
    for row in sheet.iter_rows(min_row=grp2_row, max_row=last_row, min_col=cols[0], max_col=cols[-1]):
        for cell in row:
            if is_number(cell.value) and cell.column in cols and cell.value >= 0.05:
                cell.fill = PatternFill(
                    start_color=yellow, end_color=yellow, fill_type='solid')

    wb.save(path)

def highlight_sheet_cells(path):
    '''Note: This is the customized code, not for general use'''
    wb = load_workbook(path, data_only=True)
    
    for sht_name in wb.sheetnames:
        sheet = wb[sht_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        yellow = "00FFFF00"

        grp0_row, grp1_row, grp2_row, last_row = 0, 0, 0, 0
        for i in range(max_row):
            if sheet[f'A{i+1}'].value and sheet[f'A{i+1}'].value.strip() == 'Group':
                grp0_row = i + 1
            elif sheet[f'A{i+1}'].value and sheet[f'A{i+1}'].value.strip() == 'Income':
                grp1_row = i + 1
            elif sheet[f'A{i+1}'].value and sheet[f'A{i+1}'].value.strip() == 'Expense':
                grp2_row = i + 1
            elif sheet[f'A{i+1}'].value and sheet[f'A{i+1}'].value.strip() == 'NOI':
                last_row = i + 1
                break
        cols = []   # Delta % columns
        cols_ = []  # *TD Variance columns
        for row in sheet.iter_rows(min_row=grp0_row, max_row=grp0_row, min_col=1, max_col=max_column):
            for cell in row:
                if cell.value and cell.value.strip() == 'Delta %':
                    cols.append(cell.column)
                if cell.value and 'TD Variance' in cell.value.strip():
                    cols_.append(cell.column)

        # Highlight cells in "Income" group
        for row in sheet.iter_rows(min_row=grp1_row, max_row=grp2_row-1, min_col=cols_[0], max_col=cols[-1]):
            for cell in row:
                if is_number(cell.value) and cell.column in cols and cell.value <= -0.05:
                    cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
                if is_number(cell.value) and cell.column in cols_ and cell.value <= -500:
                    cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')

        # Highlight cells in "Expense" group
        for row in sheet.iter_rows(min_row=grp2_row, max_row=last_row, min_col=cols_[0], max_col=cols[-1]):
            for cell in row:
                if is_number(cell.value) and cell.column in cols and cell.value >= 0.05:
                    cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
                if is_number(cell.value) and cell.column in cols_ and cell.value <= -500:
                    cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')

    wb.save(path)


### Note: General scripts to copy spreadsheet and keep the style format
def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)             # copy all the cell values and styles
    copy_sheet_attributes(source_sheet, target_sheet)

def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
    target_sheet._images = copy(source_sheet._images)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        pass
        # print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not only the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)